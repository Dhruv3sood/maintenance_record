import csv
import io
from typing import List, Any
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import inch
from app.models import Record
from app.utils.warranty import get_warranty_status


def export_to_csv(records: List[Record]) -> io.BytesIO:
    """Export records to CSV"""
    output = io.BytesIO()
    writer = csv.writer(output)
    
    # Header
    writer.writerow([
        "ID", "Record ID", "Created At", "Updated At",
        "Date of Delivery", "Date of Installation", "Date of Site Visit",
        "Site Visit Done By", "Installation Done By", "Commission Done By",
        "Capacity (KW)", "Heater", "Controller", "Card", "Body",
        "Client Name", "Client Phone", "Client Address", "Zone",
        "Sale Price", "Sold By", "Lead Source", "Remarks"
    ])
    
    # Data rows
    for record in records:
        writer.writerow([
            record.id,
            record.record_id,
            record.created_at.isoformat() if record.created_at else "",
            record.updated_at.isoformat() if record.updated_at else "",
            record.date_of_delivery.isoformat() if record.date_of_delivery else "",
            record.date_of_installation.isoformat() if record.date_of_installation else "",
            record.date_of_site_visit.isoformat() if record.date_of_site_visit else "",
            record.site_visit_done_by or "",
            record.installation_done_by or "",
            record.commission_done_by or "",
            record.capacity_kw or "",
            record.heater or "",
            record.controller or "",
            record.card or "",
            record.body or "",
            record.client_name,
            record.client_phone or "",
            record.client_address or "",
            record.zone or "",
            float(record.sale_price) if record.sale_price else "",
            record.sold_by or "",
            record.lead_source or "",
            record.remarks or ""
        ])
    
    output.seek(0)
    return output


def export_to_xlsx(records: List[Record]) -> io.BytesIO:
    """Export records to XLSX"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Records"
    
    # Header style
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Headers
    headers = [
        "ID", "Record ID", "Created At", "Updated At",
        "Date of Delivery", "Date of Installation", "Date of Site Visit",
        "Site Visit Done By", "Installation Done By", "Commission Done By",
        "Capacity (KW)", "Heater", "Controller", "Card", "Body",
        "Client Name", "Client Phone", "Client Address", "Zone",
        "Sale Price", "Sold By", "Lead Source", "Remarks"
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    # Data rows
    for row_idx, record in enumerate(records, 2):
        ws.cell(row=row_idx, column=1, value=record.id)
        ws.cell(row=row_idx, column=2, value=record.record_id)
        ws.cell(row=row_idx, column=3, value=record.created_at.isoformat() if record.created_at else "")
        ws.cell(row=row_idx, column=4, value=record.updated_at.isoformat() if record.updated_at else "")
        ws.cell(row=row_idx, column=5, value=record.date_of_delivery.isoformat() if record.date_of_delivery else "")
        ws.cell(row=row_idx, column=6, value=record.date_of_installation.isoformat() if record.date_of_installation else "")
        ws.cell(row=row_idx, column=7, value=record.date_of_site_visit.isoformat() if record.date_of_site_visit else "")
        ws.cell(row=row_idx, column=8, value=record.site_visit_done_by or "")
        ws.cell(row=row_idx, column=9, value=record.installation_done_by or "")
        ws.cell(row=row_idx, column=10, value=record.commission_done_by or "")
        ws.cell(row=row_idx, column=11, value=record.capacity_kw or "")
        ws.cell(row=row_idx, column=12, value=record.heater or "")
        ws.cell(row=row_idx, column=13, value=record.controller or "")
        ws.cell(row=row_idx, column=14, value=record.card or "")
        ws.cell(row=row_idx, column=15, value=record.body or "")
        ws.cell(row=row_idx, column=16, value=record.client_name)
        ws.cell(row=row_idx, column=17, value=record.client_phone or "")
        ws.cell(row=row_idx, column=18, value=record.client_address or "")
        ws.cell(row=row_idx, column=19, value=record.zone or "")
        ws.cell(row=row_idx, column=20, value=float(record.sale_price) if record.sale_price else "")
        ws.cell(row=row_idx, column=21, value=record.sold_by or "")
        ws.cell(row=row_idx, column=22, value=record.lead_source or "")
        ws.cell(row=row_idx, column=23, value=record.remarks or "")
    
    # Auto-adjust column widths
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 15
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def export_to_pdf(records: List[Record], title: str = "Records Export") -> io.BytesIO:
    """Export records to PDF"""
    output = io.BytesIO()
    doc = SimpleDocTemplate(output, pagesize=A4)
    elements = []
    
    styles = getSampleStyleSheet()
    
    # Title
    title_para = Paragraph(title, styles['Title'])
    elements.append(title_para)
    elements.append(Spacer(1, 0.2 * inch))
    
    # Prepare data
    data = [[
        "ID", "Record ID", "Delivery Date", "Client Name",
        "Zone", "Capacity", "Sale Price", "Sold By"
    ]]
    
    for record in records:
        data.append([
            str(record.id),
            record.record_id,
            record.date_of_delivery.strftime("%Y-%m-%d") if record.date_of_delivery else "",
            record.client_name[:30] if record.client_name else "",  # Truncate long names
            record.zone or "",
            record.capacity_kw or "",
            f"{float(record.sale_price):.2f}" if record.sale_price else "",
            record.sold_by or ""
        ])
    
    # Create table
    table = Table(data, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
    ]))
    
    elements.append(table)
    
    # Build PDF
    doc.build(elements)
    output.seek(0)
    return output
